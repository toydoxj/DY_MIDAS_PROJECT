import os
import csv
import json
import io
from copy import copy
from fastapi import APIRouter, Body
from fastapi.responses import StreamingResponse
from typing import List
import MIDAS_API as MIDAS

from exceptions import MidasApiError, MidasError, MidasValidationError

router = APIRouter()

_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
_FILE = os.path.join(_DATA_DIR, "floor_loads.json")
_CSV_FILE = os.path.join(_DATA_DIR, "kds_live_loads.csv")


def _read() -> list:
    if not os.path.isfile(_FILE):
        return []
    with open(_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _write(data: list):
    os.makedirs(_DATA_DIR, exist_ok=True)
    with open(_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@router.get("/kds-live-loads")
def get_kds_live_loads():
    """KDS 기준 활하중 데이터 반환 (CSV에서 읽기)"""
    if not os.path.isfile(_CSV_FILE):
        return []
    rows = []
    with open(_CSV_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({
                "category": row["대분류"],
                "detail": row["소분류"],
                "load": row["활하중"],
            })
    return rows


@router.get("/floor-loads")
def get_floor_loads():
    """저장된 Floor Load 목록 반환"""
    return _read()


@router.put("/floor-loads")
def save_floor_loads(body: List[dict] = Body(...)):
    """Floor Load 목록 전체 저장"""
    _write(body)
    return {"status": "saved", "count": len(body)}


def _find_dead_lc() -> str:
    """Static Load Case에서 Dead Load(TYPE=D) 이름 찾기"""
    try:
        raw = MIDAS.loadCaseDB.get()
        stld = raw.get("STLD", {})
        for val in stld.values():
            if isinstance(val, dict) and val.get("TYPE") == "D":
                return val.get("NAME", "Dead Load")
    except Exception:
        pass
    return "Dead Load"


def _find_live_lc() -> str:
    """Static Load Case에서 Live Load(TYPE=L) 이름 찾기"""
    try:
        raw = MIDAS.loadCaseDB.get()
        stld = raw.get("STLD", {})
        for val in stld.values():
            if isinstance(val, dict) and val.get("TYPE") == "L":
                return val.get("NAME", "Live Load")
    except Exception:
        pass
    return "Live Load"


def _build_fbld_entry(entry: dict, dead_lc: str, live_lc: str, idx: int) -> dict:
    """단일 Floor Load 엔트리를 FBLD 형식으로 변환"""
    floor = entry.get("floor", "")
    room = entry.get("roomName", "")
    name = f"{floor}_{room}" if floor and room else f"FloorLoad_{idx}"

    # 고정하중 계산: 마감하중 합계 + 슬래브 하중
    finish_total = 0.0
    for f in entry.get("finishes", []):
        v = f.get("load", "")
        if v:
            try:
                finish_total += float(v)
            except ValueError:
                pass
    slab_load = 0.0
    try:
        slab_load = float(entry.get("slabLoad", "0"))
    except ValueError:
        pass
    dead_load = finish_total + slab_load

    # 활하중
    live_load = 0.0
    try:
        live_load = float(entry.get("liveLoad", "0"))
    except ValueError:
        pass

    items = []
    if dead_load > 0:
        items.append({"LCNAME": dead_lc, "FLOOR_LOAD": -dead_load, "OPT_SUB_BEAM_WEIGHT": False})
    if live_load > 0:
        items.append({"LCNAME": live_lc, "FLOOR_LOAD": -live_load, "OPT_SUB_BEAM_WEIGHT": False})

    return {"NAME": name, "DESC": entry.get("usageDetail", ""), "ITEM": items}


@router.post("/floor-loads/sync-midas")
def sync_floor_loads_to_midas():
    """저장된 Floor Load 데이터를 MIDAS FBLD로 동기화 (NAME 기준 매칭, 없으면 신규)"""
    entries = _read()
    if not entries:
        raise MidasValidationError("동기화할 데이터가 없습니다")

    dead_lc = _find_dead_lc()
    live_lc = _find_live_lc()

    # 기존 MIDAS FBLD 데이터 가져오기
    try:
        existing_raw = MIDAS.floorLoadDB.get()
    except Exception:
        existing_raw = {}
    existing_fbld = existing_raw.get("FBLD", {})

    # 기존 데이터를 NAME → key 맵으로 변환
    name_to_key: dict[str, str] = {}
    for key, val in existing_fbld.items():
        if isinstance(val, dict) and "NAME" in val:
            name_to_key[val["NAME"]] = key

    # 다음 사용 가능한 key 번호
    max_key = max((int(k) for k in existing_fbld if k.isdigit()), default=0)

    updated_fbld = {}
    for idx, entry in enumerate(entries, start=1):
        fbld_entry = _build_fbld_entry(entry, dead_lc, live_lc, idx)
        name = fbld_entry["NAME"]

        if name in name_to_key:
            # 기존 항목 업데이트 (같은 key 유지)
            updated_fbld[name_to_key[name]] = fbld_entry
        else:
            # 새 항목 추가
            max_key += 1
            updated_fbld[str(max_key)] = fbld_entry

    try:
        MIDAS.floorLoadDB._data = {"FBLD": updated_fbld}
        MIDAS.floorLoadDB.sync()
    except Exception as e:
        raise MidasApiError("Floor Load MIDAS 동기화 실패", cause=str(e))

    return {"status": "synced", "count": len(updated_fbld)}


@router.get("/floor-loads/import-midas")
def import_from_midas():
    """MIDAS FBLD에서 현재 등록목록에 없는 항목을 가져오기"""
    try:
        raw = MIDAS.floorLoadDB.get()
    except Exception as e:
        raise MidasApiError("Floor Load MIDAS 불러오기 실패", cause=str(e))

    fbld = raw.get("FBLD", {})

    # 모든 MIDAS 항목을 가져오기 (등록목록과 비교 없이)
    # 프론트엔드에서 중복 여부를 사용자가 판단
    imported = []
    max_id = 0
    # LC이름 미리 조회 (루프 밖)
    dead_lc_name = _find_dead_lc()
    live_lc_name = _find_live_lc()

    for val in fbld.values():
        if not isinstance(val, dict):
            continue
        name = val.get("NAME", "")

        # NAME에서 층수_실명 파싱 (_가 있으면 분리, 없으면 전체를 실명으로)
        if "_" in name:
            parts = name.split("_", 1)
            floor = parts[0]
            room = parts[1]
        else:
            floor = ""
            room = name

        # ITEM에서 고정하중/활하중 추출
        dead_load = ""
        live_load = ""
        for item in val.get("ITEM", []):
            load_val = abs(item.get("FLOOR_LOAD", 0))
            lc = item.get("LCNAME", "")
            if lc == dead_lc_name:
                dead_load = str(round(load_val, 2))
            elif lc == live_lc_name:
                live_load = str(round(load_val, 2))
            elif not dead_load:
                # 첫 번째 항목을 고정하중으로, 두 번째를 활하중으로 fallback
                dead_load = str(round(load_val, 2))
            elif not live_load:
                live_load = str(round(load_val, 2))

        max_id += 1
        imported.append({
            "id": max_id,
            "floor": floor,
            "roomName": room,
            "desc": "",
            "finishes": [
                {"material": "MIDAS 불러오기", "density": "", "thickness": "", "load": dead_load},
                {"material": "", "density": "", "thickness": "", "load": ""},
                {"material": "", "density": "", "thickness": "", "load": ""},
                {"material": "", "density": "", "thickness": "", "load": ""},
                {"material": "", "density": "", "thickness": "", "load": ""},
                {"material": "", "density": "", "thickness": "", "load": ""},
            ],
            "slabType": "없음",
            "slabThickness": "",
            "slabLoad": "",
            "usageCategory": "",
            "usageDetail": val.get("DESC", ""),
            "liveLoad": live_load,
        })

    return {"imported": imported, "count": len(imported)}


@router.get("/floor-loads/export-excel")
def export_excel():
    """등록 목록을 동양구조 양식 Excel로 내보내기"""
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    entries = _read()
    if not entries:
        raise MidasValidationError("내보낼 데이터가 없습니다")

    template_path = os.path.join(_DATA_DIR, "floor_load_template.xlsx")
    if not os.path.isfile(template_path):
        raise MidasError("Excel 템플릿 파일을 찾을 수 없습니다")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 8행의 폰트/정렬/숫자포맷을 참조용으로 복사
    ref_fonts = {}
    ref_aligns = {}
    ref_numfmts = {}
    for col in range(2, 11):  # B~J
        cell = ws.cell(8, col)
        ref_fonts[col] = copy(cell.font)
        ref_aligns[col] = copy(cell.alignment)
        ref_numfmts[col] = cell.number_format

    # 샘플 데이터의 병합 셀 해제 후 행 삭제
    merges_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= 8]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))
    ws.delete_rows(8, ws.max_row - 7)

    # 프로젝트명 설정
    try:
        proj_raw = MIDAS.projectDB.get()
        pjcf = proj_raw.get("PJCF", {})
        proj_data = next(iter(pjcf.values()), {})
        project_name = proj_data.get("PROJECT", "")
        if project_name:
            ws.cell(4, 2).value = f"Project : {project_name}"
    except Exception:
        pass

    # ── 테두리/채우기 스타일 정의 ──
    thin = Side(style='thin')
    dotted = Side(style='dotted')
    no_side = Side(style=None)

    # 열별 좌/우 테두리 (템플릿 기준)
    col_sides = {
        2: (thin, dotted),    # B
        3: (dotted, thin),    # C
        4: (thin, dotted),    # D
        5: (dotted, dotted),  # E
        6: (dotted, dotted),  # F
        7: (dotted, thin),    # G
        8: (thin, thin),      # H
        9: (thin, thin),      # I
        10: (thin, thin),     # J
    }

    # 마감하중 소계: 흰색배경 25% 더 어둡게
    subtotal_fill = PatternFill(patternType='solid', fgColor='BFBFBF')

    def make_border(col, is_first, is_last):
        """실명 블록 내 위치에 따라 테두리 생성 (내부 수평 테두리 없음)"""
        left, right = col_sides.get(col, (thin, thin))
        top = thin if is_first else no_side
        bottom = thin if is_last else no_side
        return Border(left=left, right=right, top=top, bottom=bottom)

    def style_cell(cell, col, is_first=False, is_last=False, bold=False, fill=None):
        """셀 스타일 적용"""
        ref_font = ref_fonts.get(col)
        if bold and ref_font:
            cell.font = Font(name=ref_font.name, size=ref_font.size, bold=True)
        elif ref_font:
            cell.font = copy(ref_font)
        cell.border = make_border(col, is_first, is_last)
        cell.alignment = copy(ref_aligns.get(col, Alignment()))
        cell.number_format = ref_numfmts.get(col, "General")
        if fill:
            cell.fill = fill

    row = 8  # 데이터 시작 행

    for entry in entries:
        floor = entry.get("floor", "")
        room_name = entry.get("roomName", "")
        usage_detail = entry.get("usageDetail", "")
        finishes = entry.get("finishes", [])
        slab_type = entry.get("slabType", "없음")
        slab_thickness = entry.get("slabThickness", "")
        slab_load_str = entry.get("slabLoad", "")
        live_load_str = entry.get("liveLoad", "")

        # 비어있지 않은 마감재만 필터
        active_finishes = [f for f in finishes if f.get("material") or f.get("load")]
        if not active_finishes:
            active_finishes = [{"material": "", "density": "", "thickness": "", "load": ""}]

        # 항목 행 수 계산: 마감재 + 마감하중소계 + (슬래브) + 고정하중계
        has_slab = slab_type != "없음" and slab_load_str
        total_rows = len(active_finishes) + 1 + (1 if has_slab else 0) + 1
        start_row = row
        end_row = start_row + total_rows - 1

        # 마감하중 합계
        finish_total = 0.0
        for f in active_finishes:
            v = f.get("load", "")
            try:
                finish_total += float(v)
            except (ValueError, TypeError):
                pass

        slab_load = 0.0
        try:
            slab_load = float(slab_load_str)
        except (ValueError, TypeError):
            pass

        dead_load = finish_total + slab_load
        live_load = 0.0
        try:
            live_load = float(live_load_str)
        except (ValueError, TypeError):
            pass

        # ── 마감재 행들 ──
        for i, f in enumerate(active_finishes):
            is_first = (row == start_row)
            is_last = (row == end_row)
            for col in range(2, 11):
                style_cell(ws.cell(row, col), col, is_first=is_first, is_last=is_last)

            if i == 0:
                ws.cell(row, 2).value = floor        # B: 층수
                ws.cell(row, 3).value = room_name     # C: 실명
                h_cell = ws.cell(row, 8)
                h_cell.value = f"[{usage_detail}]" if usage_detail else ""  # H: 용도(활하중 근거)
                h_cell.font = Font(name=ref_fonts[8].name, size=7)

            ws.cell(row, 4).value = f.get("material", "")  # D: 재료마감
            thick = f.get("thickness", "")
            dens = f.get("density", "")
            load = f.get("load", "")
            if thick:
                try:
                    ws.cell(row, 5).value = float(thick)
                except ValueError:
                    ws.cell(row, 5).value = thick
            if dens:
                try:
                    ws.cell(row, 6).value = float(dens)
                except ValueError:
                    ws.cell(row, 6).value = dens
            if load:
                try:
                    ws.cell(row, 7).value = float(load)
                except ValueError:
                    ws.cell(row, 7).value = load
            row += 1

        # ── 마감하중 소계 행: D~G 25% 회색 배경, G열만 bold ──
        is_first = (row == start_row)
        is_last = (row == end_row)
        for col in range(2, 11):
            fill = subtotal_fill if 4 <= col <= 7 else None
            bold = (col == 7)
            style_cell(ws.cell(row, col), col, is_first=is_first, is_last=is_last,
                       bold=bold, fill=fill)
        ws.cell(row, 4).value = "마감하중 소계"
        ws.cell(row, 7).value = round(finish_total, 2)
        row += 1

        # ── 슬래브 행 ──
        if has_slab:
            is_first = (row == start_row)
            is_last = (row == end_row)
            for col in range(2, 11):
                style_cell(ws.cell(row, col), col, is_first=is_first, is_last=is_last)
            ws.cell(row, 4).value = slab_type
            if slab_thickness:
                try:
                    ws.cell(row, 5).value = float(slab_thickness)
                except ValueError:
                    pass
            # 슬래브 밀도 (타입별)
            density_map = {"철근콘크리트 슬래브": 24, "트러스형 데크슬래브": 23, "골형 데크슬래브": 18}
            d = density_map.get(slab_type)
            if d:
                ws.cell(row, 6).value = d
            ws.cell(row, 7).value = round(slab_load, 2)
            row += 1

        # ── 고정하중 계 행: D~G 흰색 배경(채우기 없음), G~J bold ──
        for col in range(2, 11):
            bold = (col >= 7)  # G~J 값 bold
            style_cell(ws.cell(row, col), col, is_first=(row == start_row),
                       is_last=True, bold=bold)
        ws.cell(row, 4).value = "고정하중 계"
        ws.cell(row, 7).value = round(dead_load, 2)
        ws.cell(row, 8).value = round(live_load, 2) if live_load else None
        ws.cell(row, 9).value = round(dead_load + live_load, 2)
        ws.cell(row, 10).value = round(1.2 * dead_load + 1.6 * live_load, 2)
        row += 1

        # ── 셀 병합: 층수(B), 실명(C), 활하중근거(H), 사용하중(I), 계수하중(J) ──
        if end_row > start_row:
            for merge_col in ["B", "C"]:
                ws.merge_cells(f"{merge_col}{start_row}:{merge_col}{end_row}")
            # H는 마감재 행 범위(고정하중계 전까지)
            h_end = end_row - 1  # 고정하중계 행 제외
            if h_end > start_row:
                ws.merge_cells(f"H{start_row}:H{h_end}")
            # I, J는 마감재+소계 범위(고정하중계 전까지)
            if h_end > start_row:
                ws.merge_cells(f"I{start_row}:I{h_end}")
                ws.merge_cells(f"J{start_row}:J{h_end}")

    # Excel을 메모리에 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # openpyxl이 누락시킨 리소스(drawing, richData, media 등)를 템플릿에서 복원
    import zipfile

    template_zip = zipfile.ZipFile(template_path, 'r')
    out_buf = io.BytesIO()

    with zipfile.ZipFile(buf, 'r') as saved_zip, \
         zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:

        saved_names = set(saved_zip.namelist())
        tmpl_names = set(template_zip.namelist())

        # openpyxl이 누락시키는 파일 목록 (drawing, media, richData, metadata)
        restore_prefixes = ('xl/drawings/', 'xl/media/', 'xl/richData/')
        restore_files = {n for n in tmpl_names
                         if any(n.startswith(p) for p in restore_prefixes)
                         or n == 'xl/metadata.xml'}

        # 저장된 파일 복사 + 패치
        import re

        for item in saved_zip.namelist():
            data = saved_zip.read(item)

            if item == '[Content_Types].xml':
                # 템플릿의 Content_Types 사용 (누락된 타입 포함)
                data = template_zip.read(item)

            elif item == 'xl/worksheets/sheet1.xml':
                content = data.decode('utf-8')
                # r: namespace 추가
                if 'xmlns:r=' not in content:
                    content = content.replace(
                        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
                        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"')
                # drawing 참조 삽입
                if '<drawing' not in content:
                    content = content.replace(
                        '</worksheet>',
                        '<drawing r:id="rId2"/></worksheet>')
                # I2 셀의 vm 속성 복원 (richData 셀 이미지 참조)
                content = re.sub(
                    r'(<c r="I2" [^>]*)(t="e")',
                    r'\1\2 vm="1"',
                    content)
                data = content.encode('utf-8')

            elif item == 'xl/_rels/workbook.xml.rels':
                # openpyxl rels에 템플릿의 richData/metadata 관계 추가
                content = data.decode('utf-8')
                tmpl_rels = template_zip.read(item).decode('utf-8')
                # 템플릿에서 richData/metadata 관계 추출
                extra_rels = re.findall(
                    r'<Relationship [^>]*(?:richdata|richvalue|rdRich|metadata)[^>]*/?>',
                    tmpl_rels, re.IGNORECASE)
                if extra_rels:
                    # 기존 최대 rId 번호 찾기
                    existing_ids = [int(x) for x in re.findall(r'Id="rId(\d+)"', content)]
                    next_id = max(existing_ids) + 1 if existing_ids else 10
                    for rel in extra_rels:
                        # rId를 충돌 없는 번호로 재할당
                        new_rel = re.sub(r'Id="rId\d+"', f'Id="rId{next_id}"', rel)
                        # Target 경로를 상대경로로 통일
                        new_rel = new_rel.replace('Target="/xl/', 'Target="')
                        content = content.replace(
                            '</Relationships>',
                            new_rel + '</Relationships>')
                        next_id += 1
                data = content.encode('utf-8')

            out_zip.writestr(item, data)

        # 템플릿에만 있는 파일 복원 (rels, drawing, richData, media, metadata)
        for tf in tmpl_names:
            if tf not in saved_names:
                out_zip.writestr(tf, template_zip.read(tf))

    template_zip.close()
    out_buf.seek(0)

    return StreamingResponse(
        out_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=floor_load_table.xlsx"},
    )
